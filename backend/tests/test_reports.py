"""Report generation: content, formats, access and audit trail (PRD A.7.8)."""

from __future__ import annotations

import io
import uuid
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.security import utcnow, today_utc
from app.models.audit import AuditLog
from app.models.employee import Employee
from app.models.enums import AuditAction, OnboardingStatus, TaskStatus
from app.models.task import EmployeeTask
from app.services.reports import FORMATS, build, generate
from tests.conftest import API
from tests.factories import make_png, upload_file

ANANYA = {
    "first_name": "Ananya",
    "last_name": "Sharma",
    "work_email": "ananya.reports@example.com",
    "department": "Engineering",
    "designation": "Backend Engineer",
}
ROHIT = {
    "first_name": "Rohit",
    "last_name": "Verma",
    "work_email": "rohit.reports@example.com",
    "department": "Finance",
    "designation": "Analyst",
}


def create_employee(client, headers, payload) -> uuid.UUID:
    response = client.post(f"{API}/employees", json=payload, headers=headers)
    assert response.status_code == 201, response.text
    return uuid.UUID(response.json()["id"])


def sheet_of(content: bytes):
    return load_workbook(io.BytesIO(content)).active


def rows_of(sheet) -> list[list]:
    """Data rows below the title block and header row."""
    all_rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    header_index = next(i for i, row in enumerate(all_rows) if row[0] == "Code")
    return all_rows[header_index + 1 :]


# --- Catalogue -------------------------------------------------------------


def test_catalogue_is_role_filtered(client, hr_headers, admin_headers):
    hr = client.get(f"{API}/reports", headers=hr_headers).json()
    admin = client.get(f"{API}/reports", headers=admin_headers).json()

    hr_keys = {r["key"] for r in hr["reports"]}
    admin_keys = {r["key"] for r in admin["reports"]}

    assert "employee_roster" in hr_keys
    # HR is never offered an export they would be refused.
    assert "audit_trail" not in hr_keys
    assert {"audit_trail", "user_accounts"} <= admin_keys
    assert {f["key"] for f in hr["formats"]} == {"xlsx", "pdf", "csv"}


def test_reports_need_hr(client, hr_headers):
    create_employee(client, hr_headers, ANANYA)
    from tests.test_dashboard_api import accept_latest_invite

    employee_headers = accept_latest_invite(client, ANANYA["work_email"])
    assert client.get(f"{API}/reports", headers=employee_headers).status_code == 403
    assert client.get(f"{API}/reports/employee_roster", headers=employee_headers).status_code == 403
    assert client.get(f"{API}/reports/employee_roster").status_code == 401


def test_admin_only_report_is_refused_to_hr(client, hr_headers):
    response = client.get(f"{API}/reports/audit_trail", headers=hr_headers)
    assert response.status_code == 403
    assert response.json()["error"]["code"] == "forbidden"


def test_unknown_report_and_format_are_rejected(client, hr_headers):
    assert client.get(f"{API}/reports/nope", headers=hr_headers).status_code == 404

    bad = client.get(
        f"{API}/reports/employee_roster", headers=hr_headers, params={"format": "docx"}
    )
    assert bad.status_code == 422
    assert "xlsx" in bad.json()["error"]["message"]


def test_inverted_date_range_is_rejected(client, admin_headers):
    response = client.get(
        f"{API}/reports/audit_trail",
        headers=admin_headers,
        params={"date_from": "2026-03-05", "date_to": "2026-03-01"},
    )
    assert response.status_code == 422


# --- Formats ---------------------------------------------------------------


@pytest.mark.parametrize("fmt", ["xlsx", "pdf", "csv"])
def test_every_format_downloads_as_an_attachment(client, hr_headers, fmt):
    create_employee(client, hr_headers, ANANYA)
    response = client.get(
        f"{API}/reports/employee_roster", headers=hr_headers, params={"format": fmt}
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(FORMATS[fmt].media_type)
    assert "attachment" in response.headers["content-disposition"]
    assert f".{fmt}" in response.headers["content-disposition"]
    assert len(response.content) > 100


def test_xlsx_is_a_readable_workbook_with_a_frozen_header(client, hr_headers):
    create_employee(client, hr_headers, ANANYA)
    content = client.get(f"{API}/reports/employee_roster", headers=hr_headers).content

    sheet = sheet_of(content)
    assert sheet["A1"].value == "Employee roster"
    assert sheet.freeze_panes is not None
    assert "Ananya Sharma" in [row[1] for row in rows_of(sheet)]


def test_pdf_starts_with_a_pdf_header(client, hr_headers):
    create_employee(client, hr_headers, ANANYA)
    content = client.get(
        f"{API}/reports/employee_roster", headers=hr_headers, params={"format": "pdf"}
    ).content
    assert content.startswith(b"%PDF-")


def test_csv_has_headers_and_a_bom_for_excel(client, hr_headers):
    create_employee(client, hr_headers, ANANYA)
    content = client.get(
        f"{API}/reports/employee_roster", headers=hr_headers, params={"format": "csv"}
    ).content

    assert content.startswith(b"\xef\xbb\xbf")
    text = content.decode("utf-8-sig")
    assert text.splitlines()[0].startswith("Code,Name,Work email")
    assert "Ananya Sharma" in text


def test_an_empty_report_still_renders_in_every_format(client, hr_headers):
    """A workspace with nothing in it must produce a file, not a 500."""
    for fmt in FORMATS:
        response = client.get(
            f"{API}/reports/employee_roster", headers=hr_headers, params={"format": fmt}
        )
        assert response.status_code == 200, fmt
        assert len(response.content) > 0


# --- Content ---------------------------------------------------------------


def test_roster_respects_the_department_filter(client, hr_headers, db):
    create_employee(client, hr_headers, ANANYA)
    create_employee(client, hr_headers, ROHIT)

    content = client.get(
        f"{API}/reports/employee_roster",
        headers=hr_headers,
        params={"format": "csv", "department": "engineering"},
    ).content.decode("utf-8-sig")

    assert "Ananya Sharma" in content
    assert "Rohit Verma" not in content

    # The filter is printed on the report, not just applied silently — a reader
    # handed the file has to know it is looking at a subset.
    dataset = build(db, "employee_roster", department="engineering")
    assert "Department: engineering" in dataset.context[0]
    assert "1 row" in dataset.context[0]


def test_status_filter_narrows_the_roster(client, hr_headers, db):
    first = create_employee(client, hr_headers, ANANYA)
    create_employee(client, hr_headers, ROHIT)
    db.get(Employee, first).onboarding_status = OnboardingStatus.COMPLETE
    db.commit()

    content = client.get(
        f"{API}/reports/employee_roster",
        headers=hr_headers,
        params={"format": "csv", "status": "complete"},
    ).content.decode("utf-8-sig")

    assert "Ananya Sharma" in content
    assert "Rohit Verma" not in content


def test_document_compliance_reports_missing_uploads(client, hr_headers, db):
    employee_id = create_employee(client, hr_headers, ANANYA)
    upload_file(client, hr_headers, employee_id, data=make_png(),
                filename="photo.png", document_type="photo")

    dataset = build(db, "document_compliance")
    row = dataset.rows[0]
    headers = dataset.headers

    assert row[headers.index("Photo")] == "extracted"
    # Never uploaded reads differently from uploaded-but-unapproved.
    assert row[headers.index("Aadhaar")] == "not uploaded"


def test_onboarding_status_counts_open_mandatory_tasks(client, hr_headers, db):
    employee_id = create_employee(client, hr_headers, ANANYA)
    db.add_all([
        EmployeeTask(employee_id=employee_id, title="Mandatory", category="task",
                     is_mandatory=True, status=TaskStatus.PENDING),
        EmployeeTask(employee_id=employee_id, title="Optional", category="task",
                     is_mandatory=False, status=TaskStatus.PENDING),
        EmployeeTask(employee_id=employee_id, title="Done", category="task",
                     is_mandatory=True, status=TaskStatus.COMPLETED),
    ])
    db.commit()

    dataset = build(db, "onboarding_status")
    row = dataset.rows[0]
    assert row[dataset.headers.index("Mandatory tasks open")] == 1


def test_task_completion_lists_the_worst_performer_first(client, hr_headers, db):
    ananya = create_employee(client, hr_headers, ANANYA)
    rohit = create_employee(client, hr_headers, ROHIT)
    db.add_all([
        EmployeeTask(employee_id=ananya, title="A", category="task",
                     status=TaskStatus.COMPLETED),
        EmployeeTask(employee_id=ananya, title="B", category="task",
                     status=TaskStatus.COMPLETED),
        EmployeeTask(employee_id=rohit, title="C", category="task",
                     status=TaskStatus.PENDING,
                     due_date=today_utc() - timedelta(days=4)),
    ])
    db.commit()

    dataset = build(db, "task_completion")
    assert dataset.rows[0][1] == "Rohit Verma"
    assert dataset.rows[0][dataset.headers.index("Complete %")] == 0
    assert dataset.rows[0][dataset.headers.index("Overdue")] == 1


def test_task_report_omits_employees_with_no_tasks(client, hr_headers, db):
    create_employee(client, hr_headers, ANANYA)
    dataset = build(db, "task_completion")

    assert dataset.rows == []
    assert any("no assigned tasks are omitted" in line for line in dataset.context)


def test_waived_tasks_count_as_closed_in_the_report(client, hr_headers, db):
    employee_id = create_employee(client, hr_headers, ANANYA)
    db.add_all([
        EmployeeTask(employee_id=employee_id, title="Waived", category="task",
                     status=TaskStatus.WAIVED, waiver_reason="Not applicable here"),
        EmployeeTask(employee_id=employee_id, title="Open", category="task",
                     status=TaskStatus.PENDING),
    ])
    db.commit()

    dataset = build(db, "task_completion")
    assert dataset.rows[0][dataset.headers.index("Complete %")] == 50


def test_user_accounts_report_lists_staff(client, admin_headers, hr_user):
    content = client.get(
        f"{API}/reports/user_accounts", headers=admin_headers, params={"format": "csv"}
    ).content.decode("utf-8-sig")

    assert "admin@example.com" in content
    assert "hr@example.com" in content
    # A password hash must never reach an export.
    assert "$2b$" not in content


def test_audit_trail_export_applies_its_filters(client, admin_headers):
    content = client.get(
        f"{API}/reports/audit_trail",
        headers=admin_headers,
        params={"format": "csv", "action": "login_success"},
    ).content.decode("utf-8-sig")

    lines = [line for line in content.splitlines()[1:] if line.strip()]
    assert lines, "the admin fixture logged in, so there is at least one entry"
    assert all("login_success" in line for line in lines)


def test_both_formats_of_one_report_agree(client, hr_headers, db):
    """The whole point of a shared dataset: the numbers cannot drift apart."""
    create_employee(client, hr_headers, ANANYA)
    create_employee(client, hr_headers, ROHIT)

    csv_content, _, _ = generate(db, key="employee_roster", format_key="csv")
    xlsx_content, _, _ = generate(db, key="employee_roster", format_key="xlsx")

    csv_names = {line.split(",")[1] for line in csv_content.decode("utf-8-sig").splitlines()[1:]}
    xlsx_names = {row[1] for row in rows_of(sheet_of(xlsx_content))}
    assert csv_names == xlsx_names == {"Ananya Sharma", "Rohit Verma"}


# --- Audit -----------------------------------------------------------------


def test_every_export_is_audited_with_its_filters(client, hr_headers, db):
    create_employee(client, hr_headers, ANANYA)
    client.get(
        f"{API}/reports/employee_roster",
        headers=hr_headers,
        params={"format": "pdf", "department": "Engineering"},
    )

    entry = db.scalar(
        select(AuditLog)
        .where(AuditLog.action == AuditAction.REPORT_EXPORTED.value)
        .order_by(AuditLog.created_at.desc())
    )
    assert entry is not None
    assert entry.entity_id == "employee_roster"
    assert entry.detail["format"] == "pdf"
    assert entry.detail["rows"] == 1
    assert entry.detail["filters"] == {"department": "Engineering"}


def test_audit_detail_omits_filters_that_were_not_supplied(client, hr_headers, db):
    client.get(f"{API}/reports/employee_roster", headers=hr_headers)

    entry = db.scalar(
        select(AuditLog).where(AuditLog.action == AuditAction.REPORT_EXPORTED.value)
    )
    assert entry.detail["filters"] == {}
