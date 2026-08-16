"""Render a Dataset as .xlsx (openpyxl).

Deliberately plain: a title block, a frozen and filtered header row, and the
data. Excel exports exist to be sorted, pivoted and pasted elsewhere, so the
sheet stays a clean rectangle rather than a formatted document — the PDF is
the presentation format.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.reports.datasets import NUMERIC, Dataset

HEADER_FILL = PatternFill("solid", fgColor="0F172A")  # slate-900
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
CONTEXT_FONT = Font(size=9, color="64748B")  # slate-500


def render(dataset: Dataset) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    # Excel rejects sheet names over 31 characters or containing []:*?/\
    sheet.title = dataset.title[:31]

    sheet.cell(row=1, column=1, value=dataset.title).font = TITLE_FONT
    row_index = 2
    for line in [
        *dataset.context,
        f"Generated {dataset.generated_at:%Y-%m-%d %H:%M UTC}",
    ]:
        sheet.cell(row=row_index, column=1, value=line).font = CONTEXT_FONT
        row_index += 1

    header_row = row_index + 1
    for column_index, column in enumerate(dataset.columns, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=column.header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="right" if column.align == NUMERIC else "left")
        sheet.column_dimensions[get_column_letter(column_index)].width = column.width

    for offset, row in enumerate(dataset.rows, start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=offset, column=column_index, value=value)

    # Freeze the header and turn on autofilter so a long export stays usable.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    if dataset.rows:
        last_column = get_column_letter(len(dataset.columns))
        sheet.auto_filter.ref = f"A{header_row}:{last_column}{header_row + len(dataset.rows)}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
